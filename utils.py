import openpyxl
import requests
import pypandoc
import json
import re

def stringEncodingFun(string):
    if not string:
        return string
    specialCharacters = r'&nbsp;|&emsp;|&ensp;|(<!--.*?-->)'
    string = re.sub(r'&#40;', '(', string)
    string = re.sub(r'&#41;', ')', string)
    string = re.sub(r'&#34;', '\"', string)
    string = re.sub(r'&#39;', '\'', string)
    string = re.sub(r'&#8194;', '', string)
    string = re.sub(r'&#8195;', '', string)
    string = re.sub(r'&#160;', '', string)
    string = re.sub(r'&#60;', '<', string)
    string = re.sub(r'&#62;', '>', string)
    string = re.sub(r'&#38;', '&', string)
    string = re.sub(r'&#34;', '\"', string)
    string = re.sub(r'&#169;', '©', string)
    string = re.sub(r'&#174;', '®', string)
    string = re.sub(r'&#8482;', '™', string)
    string = re.sub(r'&#215;', '×', string)
    string = re.sub(r'&#247;', '÷', string)
    string = re.sub(r'&nbsp;', ' ', string)
    string = re.sub(r'&emsp;', '', string)
    string = re.sub(r'&ensp;', '', string)
    string = re.sub(r'&lt;', '<', string)
    string = re.sub(r'&gt;', '>', string)
    string = re.sub(r'&amp;', '&', string)
    string = re.sub(r'&quot;', '\"', string)
    string = re.sub(r'&copy;', '©', string)
    string = re.sub(r'&reg;', '®', string)
    string = re.sub(r'&times;', '×', string)
    string = re.sub(r'&divide;', '÷', string)
    string = re.sub(r'(<!--.*?-->)', '', string)
    string = re.sub(r'\\t', '', string)
    string = re.sub(r'\'', '"', string)
    return string


def format_html(data, file_name):
  # append_data = {}
  # if "title" not in data:
  #     data['title'] = []
  # title = data['title']
  # append_data['title'] = title
  # if "keywords" not in data:
  #     data['keywords'] = []
  # keywords = data['keywords']
  # append_data['keywords'] = keywords
  # if "description" not in data:
  #     data['description'] = []
  # description = data['description']
  # append_data['description'] = description
  # if "content" not in data:
  #     return
  # append_data = json.dumps(append_data)
  # content = append_data + '<br>' +data['content']
#   data_type = data['data_type']
#   data_second_type = data['data_second_type']
  # Create a new document


  # 获取HTML内容
  html = data
  pypandoc.convert_text(html, 'docx', 'html', outputfile=file_name)  # 将 html 代码转化成docx

def read_elsx(file_path):
  # 打开Excel文件
  workbook = openpyxl.load_workbook(file_path)

  # 获取工作表名称
  sheet_name = workbook.sheetnames[0]

  # 获取工作表对象
  sheet = workbook[sheet_name]

  # 定义一个空列表
  data = []

  # 获取标题行的值，用于作为字典的键
  headers = [cell.value for cell in sheet[1]]

  # 遍历每一行，并将每一行的数据整理成字典形式，添加到列表中
  for row in sheet.iter_rows(min_row=2):
      # 定义一个空字典，用于存储当前行的数据
      row_data = {}
      for index, cell in enumerate(row):
          # 将单元格的值添加到当前行的数据字典中
          row_data[headers[index]] = cell.value
      # 将当前行的数据字典添加到整个数据列表中
      data.append(row_data)

  # 打印整个数据列表
  return data


